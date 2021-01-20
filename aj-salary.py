# -*- coding: UTF-8 -*-
import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
import os
import const_define
import re
from decimal import Decimal, ROUND_HALF_UP
from collections import OrderedDict
import shutil
import io
# import time


class PerformanceCalculation(object):
    def __init__(self):
        self.salary_folder_filename = const_define.SALARY_DATA_FOLDER_NAME
        self.sheet = None
        self.sheet_formula = None
        self.sheet_statistical = None
        self.sheet_statistical_formula = None
        self.name_mapping_dic = OrderedDict()
        self.personal_bonus_dic = {}
        self.overall_bonus_dic = {}
        self.total_bonus_dic = {}

        self.get_name_mapping_dic()

    def get_name_mapping_dic(self):
        with open('name.db', 'r', encoding='UTF-8') as file_h:
            lines = file_h.readlines()
            for line in lines:
                line_name = line.split(':')
                chinese_name = line_name[0].rstrip('\n\r')
                english_name = line_name[1].rstrip('\n\r')
                self.name_mapping_dic[english_name] = chinese_name

    def calc_all_salary(self):
        performance_files = self.get_all_perf_file_name()
        self.get_statistical_table_dic()
        # print(self.overall_bonus_dic)
        for full_filepath in performance_files:
            filename_full = os.path.basename(full_filepath)
            filename_with_ext = os.path.splitext(filename_full)
            filename = filename_with_ext[0]
            file_ext = filename_with_ext[1]

            if filename.find(const_define.STATISTICAL_TABLE_KEYWORD) < 0 and file_ext == '.xlsx':
                name = self.get_name_from_filename(filename).lower()
                print(name)
                personal_bonus = self.calc_salary(full_filepath)
                self.personal_bonus_dic[name] = personal_bonus
                print('個人獎金: {}'.format(personal_bonus))
                print('統籌獎金: {}'.format(self.overall_bonus_dic[name]))
                print('------------------------------------------------')
                total_bonus = self.overall_bonus_dic[name] + int(personal_bonus)

                self.total_bonus_dic[name] = total_bonus

        for name in self.name_mapping_dic.keys():
            if name not in self.total_bonus_dic.keys():
                continue
            print(name)
            print('業績獎金: {}'.format(self.total_bonus_dic[name]))
            print('======================')

        self.move_file_to_backup_folder(performance_files)

    def get_statistical_table_dic(self):
        statistical_table_filename = self.get_statistical_table_filename(self.get_all_perf_file_name())

        with open(statistical_table_filename, "rb") as f:
            in_mem_file = io.BytesIO(f.read())
        workbook = openpyxl.load_workbook(in_mem_file, read_only=True, data_only=False)
        worksheets = tuple(workbook.sheetnames)
        self.sheet_statistical_formula = workbook[worksheets[0]]

        with open(statistical_table_filename, "rb") as f:
            in_mem_file2 = io.BytesIO(f.read())
        workbook_temp = openpyxl.load_workbook(in_mem_file2, read_only=True, data_only=True)
        worksheets_temp = tuple(workbook_temp.sheetnames)
        self.sheet_statistical = workbook_temp[worksheets_temp[0]]

        for row in self.sheet_statistical_formula.iter_rows():
            for cell in row:
                if cell.value == '各別獎金':
                    self.get_overall_value(cell)
                    return

    def get_overall_value(self, cell):
        # for name in self.name_mapping_dic.keys():
        for i in range(1, 99):
            name = self.sheet_statistical_formula.cell(row=cell.row-1, column=cell.column+i)
            name = name.value
            if name:
                for key in self.name_mapping_dic.keys():
                    if key == name.lower():
                        cell_value = self.sheet_statistical_formula.cell(row=cell.row+1, column=cell.column+i).value
                        # value = self.round_v2(value)
                        cell_value = re.sub("=", "", cell_value)
                        # print(self.sheet_statistical_formula[cell_value].value)
                        overall_value = self.calculate_value_cell(self.sheet_statistical_formula[cell_value], self.sheet_statistical)
                        # print(overall_value)
                        self.overall_bonus_dic[key] = overall_value
            else:
                return

    def calc_salary(self, full_filename):
        with open(full_filename, "rb") as f:
            in_mem_file = io.BytesIO(f.read())
        workbook = openpyxl.load_workbook(in_mem_file, read_only=True, data_only=True)
        worksheets = tuple(workbook.sheetnames)
        self.sheet = workbook[worksheets[0]]

        with open(full_filename, "rb") as f:
            in_mem_file2 = io.BytesIO(f.read())
        workbook_temp = openpyxl.load_workbook(in_mem_file2, read_only=True, data_only=False)
        worksheets_temp = tuple(workbook_temp.sheetnames)
        self.sheet_formula = workbook_temp[worksheets_temp[0]]

        person_col = column_index_from_string('I')

        # start = time.time()
        for row in self.sheet_formula.iter_rows(min_col=person_col, max_col=person_col, min_row=300):
            for cell in row:
                if cell.value == '個人Total' and cell.column == person_col:
                    perf_value = self.calculate_value_cell(self.sheet_formula.cell
                                                           (row=cell.row, column=cell.column+1), self.sheet)
                    # end = time.time()
                    # print("執行時間：%f 秒" % (end - start))
                    return perf_value

    def calculate_value_cell(self, cell, sheet_h):
        formula = cell.value
        performance_value = 0

        if not self.is_formula(formula):
            return formula

        if formula.find("SUM") >= 0:
            sum_index = formula.split(':')
            sum_min_index = coordinate_from_string(re.sub(r'=SUM\(', '', sum_index[0]))
            sum_max_index = coordinate_from_string(re.sub(r'\)', '', sum_index[1]))
            # print(sum_min_index)
            # print(sum_max_index)
            sum_min_col = column_index_from_string(sum_min_index[0])
            sum_min_row = int(sum_min_index[1])
            sum_max_col = column_index_from_string(sum_max_index[0])
            sum_max_row = int(sum_max_index[1])

            if sum_min_col != sum_max_col:
                print('not sum the same col, not support at this time!')
                return
            # row_count = sum_min_row
            # print(sum_min_col, sum_min_row, sum_max_col, sum_max_row)
            # print(row_count)

            for row in sheet_h.iter_rows(min_col=sum_min_col, max_col=sum_max_col,
                                           min_row=sum_min_row, max_row=sum_max_row):
                for cell in row:
                    value = cell.value
                    if value:
                        try:
                            value = self.round_v2(value)
                            performance_value = performance_value + value
                            # print(value)
                        except Exception as e:
                            str(e)
            return performance_value
        elif formula.find("+") >= 0:
            formula = re.sub('=', '', formula)
            cell_list = formula.split('+')
            for cell_index in cell_list:
                value = sheet_h[cell_index].value

                if value:
                    # print(value)
                    try:
                        float(value)
                        # print(value)
                        value = self.round_v2(value)
                        # print(value)
                        performance_value = performance_value + value
                    except Exception as e:
                        str(e)
            return performance_value
        else:
            print('not SUM formula')

    @staticmethod
    def round_v2(number):
        origin_num = Decimal(number)
        result_num = origin_num.quantize(Decimal('0'), rounding=ROUND_HALF_UP)
        return result_num

    @staticmethod
    def is_formula(string):
        try:
            if string[0] == '=':
                return True
            else:
                return False
        except Exception as e:
            str(e)
            return False

    @staticmethod
    def get_name_from_filename(filename):
        temp_str = filename.split(" ")[0]
        name = temp_str.split("_")[1]
        return name

    @staticmethod
    def get_statistical_table_filename(filelist):
        for full_filename in filelist:
            file_name_ext = os.path.basename(full_filename)
            filename = os.path.splitext(file_name_ext)[0]
            file_ext = os.path.splitext(file_name_ext)[1]
            if filename.find(const_define.STATISTICAL_TABLE_KEYWORD) > -1 and file_ext == '.xlsx':
                return full_filename
        return None

    def get_all_perf_file_name(self):
        file_list = []
        for file_name in os.listdir(self.salary_folder_filename):
            full_path = os.path.join(self.salary_folder_filename, file_name)
            if os.path.isfile(full_path):
                file_list.append(full_path)
        return file_list

    @staticmethod
    def move_file_to_backup_folder(file_list):
        if not os.path.isdir(const_define.BACKUP_FOLDER_PATH):
            os.mkdir(const_define.BACKUP_FOLDER_PATH)

        backup_mon_folder_name = os.path.join(const_define.BACKUP_FOLDER_PATH,
                                              re.findall('.*月', os.path.basename(file_list[1]))[0])
        if not os.path.isdir(backup_mon_folder_name):
            os.mkdir(backup_mon_folder_name)

        for file_path in file_list:
            shutil.move(file_path, backup_mon_folder_name)


salary_h = PerformanceCalculation()
salary_h.calc_all_salary()
print("")
input('請Enter鍵離開...')
