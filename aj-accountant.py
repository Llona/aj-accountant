# -*- coding: UTF-8 -*-
import openpyxl
import shutil
import os
import const_define


# active sheet name
# print(workbook.active)

# load excel file
workbook = openpyxl.load_workbook(const_define.DETAILED_LEDGER_FULL_PATH, data_only=True)
# get all sheet name
# worksheets = workbook.get_sheet_names()
worksheets = tuple(workbook.sheetnames)
print(worksheets)

for i in worksheets:
    print(i)

# get sheet content
sheet = workbook[worksheets[0]]
# print(sheet)
# print(sheet.title)
# print(sheet.cell(row=2, column=2).value)
# print(sheet['J2'].value)

for rowOfCell in sheet['B2':'M2']:
    for cell in rowOfCell:
        # print(cell.coordinate, cell.value)
        print(cell.value)

# shutil.copyfile(os.path.join('T:'), os.path.join('ttt.txt'))
# filename_netdriver = os.path.join(r"t:", 'vv')
# filename_netdriver = os.path.join(filename_netdriver, 'Roy')
# filename_netdriver = os.path.join(filename_netdriver, 'command.txt')
# print(filename_netdriver)
# shutil.copy(filename_netdriver, DATA_FOLDER_FULL_PATH)

# filename_netdriver = os.path.join(r"t:", r'vv\Roy\command.txt')
# print(filename_netdriver)
# shutil.copy(filename_netdriver, DATA_FOLDER_FULL_PATH)
