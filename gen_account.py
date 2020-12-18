# -*- coding: UTF-8 -*-
import openpyxl
import os

# load excel file
FOLDER_PATH = os.path.join(os.getcwd(), 'data')
ACCOUNT_DATA_FULL_PATH = os.path.join(FOLDER_PATH, r'匯款資料準備.xlsx')
workbook = openpyxl.load_workbook(ACCOUNT_DATA_FULL_PATH, data_only=True)
# get all sheet name
# worksheets = workbook.get_sheet_names()
worksheets = tuple(workbook.sheetnames)
# print(worksheets)
#
# for i in worksheets:
#     print(i)

# get sheet content
sheet = workbook[worksheets[0]]

# print(sheet)
# print(sheet.title)
# print(sheet.cell(row=2, column=2).value)
# print(sheet['J2'].value)
print(sheet.max_row, sheet.max_column)
print(sheet.cell(row=sheet.max_row, column=2).value)
